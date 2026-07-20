"""
Sephora Brands tool for the Yearly Asana Task platform.

Mirrors the existing Ulta "Top Beauty Brands" tool: an A-Z directory of every
brand carried on https://www.sephora.com/brands-list, each with its Sephora
brand page plus auto-generated social handles (Facebook / Instagram / X /
YouTube / TikTok) and a Wikipedia link where one is known.

Data model
----------
Brand names come from `sephora_brands.json` if that file sits next to this
module (this is what `fetch_sephora_brands.py` writes when run locally). If the
file is absent, a built-in starter snapshot (SEED_BRANDS) is used so the page
renders out of the box. Social handles are slugified from the brand name and are
UNVERIFIED unless present in VERIFIED_HANDLES -- exactly the same convention the
Ulta tool uses for ranks 11-500.

Integration
-----------
    from sephora_brands import router as sephora_router
    app.include_router(sephora_router)

That exposes:
    GET /sephora-brands
    GET /sephora-brands/export?fmt=csv
    GET /sephora-brands/export?fmt=xlsx
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import unicodedata

from fastapi import APIRouter, Query
from fastapi.responses import HTMLResponse, StreamingResponse

router = APIRouter()

_HERE = os.path.dirname(os.path.abspath(__file__))
_DATA_FILE = os.path.join(_HERE, "sephora_brands.json")

# --------------------------------------------------------------------------- #
# Handle overrides for well-known brands whose real social handle or Sephora
# page slug is NOT what a naive slug of the display name would produce.
# {display_name: {"page": sephora-slug, "handle": social-slug, "wiki": bool}}
# Any field may be omitted; missing fields fall back to the derived slug.
# --------------------------------------------------------------------------- #
VERIFIED_HANDLES: dict[str, dict] = {
    "Fenty Beauty by Rihanna": {"page": "fenty-beauty", "handle": "fentybeauty", "wiki": True},
    "Fenty Skin": {"page": "fenty-skin", "handle": "fentyskin"},
    "Rare Beauty by Selena Gomez": {"page": "rare-beauty", "handle": "rarebeauty", "wiki": True},
    "Rhode": {"page": "rhode", "handle": "rhode"},
    "Haus Labs by Lady Gaga": {"page": "haus-labs-by-lady-gaga", "handle": "hauslabs"},
    "One/Size by Patrick Starrr": {"page": "onesize-beauty", "handle": "onesizebeauty"},
    "NARS": {"page": "nars", "handle": "narsissist", "wiki": True},
    "Yves Saint Laurent": {"page": "yves-saint-laurent", "handle": "yslbeauty", "wiki": True},
    "Armani Beauty": {"page": "giorgio-armani-beauty", "handle": "armanibeauty"},
    "Dior": {"page": "dior", "handle": "dior", "wiki": True},
    "Prada Beauty": {"page": "prada-beauty", "handle": "pradabeauty"},
    "Valentino Beauty": {"page": "valentino-beauty", "handle": "maisonvalentino"},
    "The Ordinary": {"page": "the-ordinary", "handle": "theordinary", "wiki": True},
    "The INKEY List": {"page": "the-inkey-list", "handle": "theinkeylist"},
    "Charlotte Tilbury": {"page": "charlotte-tilbury", "handle": "charlottetilbury", "wiki": True},
    "Anastasia Beverly Hills": {"page": "anastasia-beverly-hills", "handle": "anastasiabeverlyhills", "wiki": True},
    "Huda Beauty": {"page": "huda-beauty", "handle": "hudabeauty", "wiki": True},
    "Sol de Janeiro": {"page": "sol-de-janeiro", "handle": "soldejaneiro", "wiki": True},
    "Drunk Elephant": {"page": "drunk-elephant", "handle": "drunkelephant", "wiki": True},
    "Sephora Collection": {"page": "sephora-collection", "handle": "sephoracollection", "wiki": True},
    "La Roche-Posay": {"page": "la-roche-posay", "handle": "larocheposayusa", "wiki": True},
    "Clinique": {"page": "clinique", "handle": "clinique", "wiki": True},
    "Tarte": {"page": "tarte", "handle": "tartecosmetics", "wiki": True},
    "Too Faced": {"page": "too-faced", "handle": "toofaced", "wiki": True},
    "Urban Decay": {"page": "urban-decay", "handle": "urbandecaycosmetics", "wiki": True},
    "Benefit Cosmetics": {"page": "benefit-cosmetics", "handle": "benefitcosmetics", "wiki": True},
    "Fresh": {"page": "fresh", "handle": "fresh", "wiki": True},
    "Tatcha": {"page": "tatcha", "handle": "tatcha", "wiki": True},
    "Guerlain": {"page": "guerlain", "handle": "guerlain", "wiki": True},
    "Shiseido": {"page": "shiseido", "handle": "shiseido", "wiki": True},
    "Olaplex": {"page": "olaplex", "handle": "olaplex", "wiki": True},
    "Moroccanoil": {"page": "moroccanoil", "handle": "moroccanoil", "wiki": True},
    "Bobbi Brown": {"page": "bobbi-brown", "handle": "bobbibrown", "wiki": True},
    "Kayali": {"page": "kayali", "handle": "kayali"},
    "Rabanne": {"page": "rabanne", "handle": "rabanne", "wiki": True},
    "Gucci": {"page": "gucci", "handle": "gucci", "wiki": True},
    "MAC Cosmetics": {"page": "mac-cosmetics", "handle": "maccosmetics", "wiki": True},
    "Estée Lauder": {"page": "estee-lauder", "handle": "esteelauder", "wiki": True},
    "Lancôme": {"page": "lancome", "handle": "lancomeofficial", "wiki": True},
    "La Mer": {"page": "la-mer", "handle": "lamer", "wiki": True},
    "Kérastase": {"page": "kerastase", "handle": "kerastase_official", "wiki": True},
    "Kiehl's Since 1851": {"page": "kiehls", "handle": "kiehls", "wiki": True},
    "L'Occitane": {"page": "l-occitane", "handle": "loccitane", "wiki": True},
    "MAKE UP FOR EVER": {"page": "make-up-for-ever", "handle": "makeupforeverofficial", "wiki": True},
    "TOM FORD": {"page": "tom-ford", "handle": "tomford", "wiki": True},
    "Marc Jacobs Beauty": {"page": "marc-jacobs-beauty", "handle": "marcjacobsbeauty"},
    "Jo Malone London": {"page": "jo-malone-london", "handle": "jomalonelondon", "wiki": True},
    "Dolce&Gabbana": {"page": "dolce-and-gabbana", "handle": "dolcegabbana", "wiki": True},
    "shu uemura": {"page": "shu-uemura", "handle": "shuuemura", "wiki": True},
    "CHANEL": {"page": "chanel", "handle": "chanel.beauty", "wiki": True},
    "Glossier": {"page": "glossier", "handle": "glossier", "wiki": True},
    "Sunday Riley": {"page": "sunday-riley", "handle": "sundayriley", "wiki": True},
}

# --------------------------------------------------------------------------- #
# Starter snapshot. Replace/extend by running fetch_sephora_brands.py locally,
# which writes sephora_brands.json next to this file (that file wins if present).
# --------------------------------------------------------------------------- #
SEED_BRANDS: list[str] = [
    "5 SENS", "54 Thrones", "AAVRANI", "Act+Acre",
    "adwoa beauty", "AESTURA", "AKT London", "ALPYN",
    "ALTERNA Haircare", "Ami Colé", "amika", "Anastasia Beverly Hills",
    "AORA México", "Arencia", "Ariana Grande", "Armani Beauty",
    "Augustinus Bader", "Azzaro", "BaBylissPRO", "banu",
    "bareMinerals", "BASMA", "Beauty of Joseon", "BeautyBio",
    "Beautyblender", "belif", "Benefit Cosmetics", "Bio Ionic",
    "Biodance", "Biossance", "Bobbi Brown", "BondiBoost",
    "BORNTOSTANDOUT", "Bounce Curl", "Boy Smells", "BREAD BEAUTY SUPPLY",
    "Briogeo", "BROWN GIRL Jane", "Bumble and bumble", "BURBERRY",
    "By Rosie Jane", "caliray", "CANOPY", "Carolina Herrera",
    "Caudalie", "CAY SKIN", "Ceremonia", "CHANEL",
    "Charlotte Tilbury", "Chloé", "Chris McMillan", "Chunks",
    "ciele", "Cinema Secrets", "Clarins", "CLEAN RESERVE",
    "CLINIQUE", "COACH", "COLOR WOW", "Commodity",
    "Community Sixty-Six", "COOLA", "Crown Affair", "Curlsmith",
    "CYKLAR", "dae", "DAMDAM", "Danessa Myricks Beauty",
    "Davines", "DedCool", "DEREK LAM 10 CROSBY", "DERMAFLASH",
    "Dermalogica", "Dezi Skin", "Dieux", "DIOR",
    "Dolce Glow", "Dolce&Gabbana", "DOMINIQUE COSMETICS", "Donna Karan",
    "Dr. Barbara Sturm", "Dr. Dennis Gross Skincare", "Dr. Groot", "Dr. Idriss",
    "Dr. Jart+", "Dr.FORHAIR", "Drunk Elephant", "Drybar",
    "DUO", "Dyson", "EADEM", "Ebb Ocean Club",
    "EILISH FRAGRANCES", "Element Eight", "Elemis", "Ellis Brooklyn",
    "Emi Jay", "Erborian", "Estée Lauder", "Evereden",
    "Experiment", "Fable & Mane", "Facile", "Fara Homidi",
    "Farmacy", "fel beauty", "Fenty Beauty by Rihanna", "Fig.1",
    "First Aid Beauty", "Flora + Bast", "FlutterHabit", "FOREO",
    "FORVR MOOD", "Freck Beauty", "fresh", "FROMLABS",
    "FUGAZZI", "ghd", "Gisou", "Givenchy",
    "Glamnetic", "Glossier", "Glow Recipe", "Grande Cosmetics",
    "Gucci", "GUERLAIN", "Hanni", "Hanyul",
    "Harlem Perfume Co.", "HAUS LABS BY LADY GAGA", "Hello Sunday", "Henry Rose",
    "HERMÈS", "HigherDOSE", "Hourglass", "House of Lashes",
    "HUDA BEAUTY", "Hugo Boss", "HUNG VANNGO BEAUTY", "Hyper Skin",
    "Iconic London", "IGK", "ILIA", "indē wild",
    "INFLUXIOUS", "INNBEAUTY PROJECT", "innisfree", "IOPE",
    "Iris&Romeo", "ISAMAYA", "ISDIN", "Isle of Paradise",
    "IT Cosmetics", "Jack Black", "Jean Paul Gaultier", "JIMMY CHOO",
    "Jo Malone London", "JOSIE MARAN", "Juicy Couture", "Juliette Has a Gun",
    "JVN", "K18 Biomimetic Hairscience", "Kaja", "Kate McLeod",
    "Kate Somerville", "Katini Skin", "KAYALI", "Kenra Professional",
    "Kérastase", "Kiehl's Since 1851", "KILIAN Paris", "KORA Organics",
    "KORRES", "Kosas", "Kulfi", "KVD Beauty",
    "L'Occitane", "L'Oréal Professionnel", "L'Oréal Professionnel Steampod", "La Mer",
    "Laka", "Lancôme", "LANEIGE", "Laura Geller",
    "Laura Mercier", "LAWLESS", "Lilly Lashes", "Lion Pose",
    "Living Proof", "LIXR Beauty", "LORE", "LoveShackFancy",
    "Luna Daily", "Lux Unfiltered", "LYS Beauty", "m.ph by Mary Phillips",
    "MAC Cosmetics", "MAED", "Maison Louis Marie", "Maison Margiela",
    "MAKE UP FOR EVER", "MAKEUP BY MARIO", "Mane", "Mango People",
    "manucurist", "MARA", "Marc Jacobs Beauty", "Mario Badescu",
    "MATTER OF FACT", "maude", "Medik8", "Melanin Haircare",
    "Melt Cosmetics", "MERIT", "Messy by Alli Webb", "Milk Makeup",
    "Miu Miu", "Mizani", "Montale", "Montblanc",
    "Moon Juice", "Moroccanoil", "Mother Science", "Mugler",
    "Murad", "NARS", "NATASHA DENONA", "Naturally Serious",
    "Nécessaire", "NEST New York", "Nette", "NUDESTIX",
    "NuFACE", "Ogee", "Olaplex", "OLEHENRIKSEN",
    "OLIVIAUMMA", "ONE/SIZE by Patrick Starrr", "Oribe", "Origins",
    "OUAI", "OUI the People", "Ouidad", "PAT McGRATH LABS",
    "PATRICK TA", "PATTERN by Tracee Ellis Ross", "Paul Mitchell", "Paula's Choice",
    "Peace Out", "Peter Thomas Roth", "PHLUR", "PHYLA",
    "Prada", "Pureology", "Rabanne", "Rahua",
    "Ralph Lauren", "RANAVAT", "Range Beauty", "Rare Beauty by Selena Gomez",
    "REFY", "Rejuran", "rhode", "RIES",
    "ROSE Ingleton MD", "Rosebud Perfume Co.", "Rossano Ferretti Parma", "RŌZ",
    "RUHVEDA", "Ruka Hair", "Saie", "Saint Jane",
    "Salt & Stone", "Sarah Creal", "SEPHORA COLLECTION", "Sephora Favorites",
    "SEPHORA The Merch Shop", "Shani Darden Skin Care", "Shark Beauty", "SHAZ & KIKS",
    "Shiseido", "shu uemura", "Sienna Naturals", "SIMIHAZE BEAUTY",
    "Sincerely Yours", "SK-II", "Skinfix", "Skylar",
    "Slip", "Smile Makers", "SOFIE PAVITT FACE", "Soft Services",
    "Sol de Janeiro", "St. Tropez", "stila", "Stripes",
    "StriVectin", "Sulwhasoo", "Summer Fridays", "SUNDAY II SUNDAY",
    "Sunday Riley", "Supergoop!", "superzero", "T3",
    "Tabu", "tarte", "Tata Harper", "Tatcha",
    "The 7 Virtues", "The INKEY List", "The Maker", "The Ordinary",
    "The Original MakeUp Eraser", "THE STEAM BAR", "Then I Met You", "Therabody",
    "TOM FORD", "Too Faced", "Topicals", "Torriden",
    "Touchland", "Tower 28 Beauty", "TWEEZERMAN", "U Beauty",
    "Ultra Violette", "UNITE Hair", "UNOVE", "Urban Decay",
    "Vacation", "Valentino", "Vegamour", "Velour Lashes",
    "Verb", "Versace", "Viktor&Rolf", "VIOLETTE_FR",
    "Viori", "Virtue", "VOLUSPA", "Wander Beauty",
    "Westman Atelier", "Whipped", "Wonderskin", "World of Chris Collins",
    "YISE Beauty", "Youth To The People", "Yves Saint Laurent",
]

DATA_SOURCE_LABEL = "Cached snapshot"
DATA_SOURCE_URL = "https://www.sephora.com/brands-list"


# --------------------------------------------------------------------------- #
# Slug helpers
# --------------------------------------------------------------------------- #
def _ascii(name: str) -> str:
    """Strip accents/diacritics -> plain ASCII."""
    return "".join(
        c for c in unicodedata.normalize("NFKD", name) if not unicodedata.combining(c)
    )


def page_slug(name: str) -> str:
    """Sephora brand-page slug, e.g. 'La Roche-Posay' -> 'la-roche-posay'."""
    s = _ascii(name).lower()
    s = s.replace("&", " and ").replace("+", " plus ").replace("/", " ")
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def social_slug(name: str) -> str:
    """Compact social handle, e.g. 'La Roche-Posay' -> 'larocheposay'."""
    s = _ascii(name).lower()
    s = s.replace("&", "and").replace("+", "plus")
    return re.sub(r"[^a-z0-9]+", "", s)


def _brand_names() -> list[str]:
    """Load brand names from sephora_brands.json if present, else the seed."""
    if os.path.exists(_DATA_FILE):
        try:
            with open(_DATA_FILE, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            names = data.get("brands", data) if isinstance(data, dict) else data
            names = [n["name"] if isinstance(n, dict) else str(n) for n in names]
            names = [n.strip() for n in names if n and str(n).strip()]
            if names:
                # de-dupe (case-insensitive), keep first spelling, sort A-Z
                seen, out = set(), []
                for n in names:
                    k = n.lower()
                    if k not in seen:
                        seen.add(k)
                        out.append(n)
                return sorted(out, key=lambda x: _ascii(x).lower())
        except Exception:
            pass
    return sorted(SEED_BRANDS, key=lambda x: _ascii(x).lower())


def build_rows() -> list[dict]:
    """Assemble the full table: one dict per brand."""
    rows = []
    for i, name in enumerate(_brand_names(), start=1):
        ov = VERIFIED_HANDLES.get(name, {})
        pslug = ov.get("page") or page_slug(name)
        hslug = ov.get("handle") or social_slug(name)
        rows.append(
            {
                "num": i,
                "brand": name,
                "sephora_page": f"https://www.sephora.com/brand/{pslug}",
                "facebook": f"https://www.facebook.com/{hslug}",
                "instagram": f"https://www.instagram.com/{hslug}",
                "twitter": f"https://x.com/{hslug}",
                "youtube": f"https://www.youtube.com/@{hslug}",
                "tiktok": f"https://www.tiktok.com/@{hslug}",
                "wikipedia": (
                    f"https://en.wikipedia.org/wiki/{_ascii(name).replace(' ', '_')}"
                    if ov.get("wiki")
                    else ""
                ),
                "verified": name in VERIFIED_HANDLES,
            }
        )
    return rows


# --------------------------------------------------------------------------- #
# Exports
# --------------------------------------------------------------------------- #
EXPORT_HEADERS = [
    "#", "Brand", "Sephora Page", "Facebook", "Instagram",
    "X / Twitter", "YouTube", "TikTok", "Wikipedia", "Socials Verified",
]


def _row_values(r: dict) -> list:
    return [
        r["num"], r["brand"], r["sephora_page"], r["facebook"], r["instagram"],
        r["twitter"], r["youtube"], r["tiktok"], r["wikipedia"],
        "yes" if r["verified"] else "no",
    ]


def export_csv() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(EXPORT_HEADERS)
    for r in build_rows():
        w.writerow(_row_values(r))
    return buf.getvalue().encode("utf-8-sig")


def export_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Sephora Brands"
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(EXPORT_HEADERS)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")
    for r in build_rows():
        ws.append(_row_values(r))
    widths = [6, 34, 46, 40, 40, 34, 40, 40, 40, 16]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wdt
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------- #
# HTML rendering  (self-contained; matches the look of the platform's pages)
# --------------------------------------------------------------------------- #
def _cell_link(url: str, label: str, title: str) -> str:
    if not url:
        return '<td class="c">—</td>'
    return f'<td class="c"><a href="{url}" title="{title}" target="_blank" rel="noopener">{label}</a></td>'


def render_page() -> str:
    rows = build_rows()
    total = len(rows)
    verified = sum(1 for r in rows if r["verified"])
    is_snapshot = not os.path.exists(_DATA_FILE)
    snapshot_note = (
        "starter snapshot" if is_snapshot else "scraped snapshot"
    )

    body = []
    for r in rows:
        body.append(
            "<tr>"
            f'<td class="num">{r["num"]}</td>'
            f'<td class="brand">{r["brand"]}</td>'
            + _cell_link(
                r["sephora_page"],
                r["sephora_page"].replace("https://www.", ""),
                "Sephora page",
            )
            + _cell_link(r["facebook"], "Link", "Facebook")
            + _cell_link(r["instagram"], "Link", "Instagram")
            + _cell_link(r["twitter"], "Link", "X / Twitter")
            + _cell_link(r["youtube"], "Link", "YouTube")
            + _cell_link(r["tiktok"], "Link", "TikTok")
            + _cell_link(r["wikipedia"], "Wiki", "Wikipedia")
            + "</tr>"
        )
    rows_html = "\n".join(body)

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Sephora Brands · Yearly Asana task</title>
<style>
  :root {{
    --navy: #1a3a5c; --band: #dce8f5; --body: #111111;
    --line: #e6e9ee; --alt: #f7f9fc; --link: #1a5fb4;
  }}
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI",
         Roboto, Helvetica, Arial, sans-serif; color: var(--body);
         background: #fff; line-height: 1.45; }}
  .wrap {{ max-width: 1200px; margin: 0 auto; padding: 28px 20px 60px; }}
  a {{ color: var(--link); text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .back {{ font-size: 14px; }}
  .kicker {{ color: var(--navy); font-weight: 700; letter-spacing: .04em;
            text-transform: uppercase; font-size: 12px; margin-top: 14px; }}
  h1 {{ margin: 6px 0 10px; font-size: 30px; color: var(--navy); }}
  .desc {{ max-width: 820px; color: #333; }}
  .source {{ background: var(--band); color: var(--navy); border-radius: 8px;
            padding: 10px 14px; font-size: 13px; margin: 16px 0; }}
  .note {{ font-size: 13px; color: #555; margin: 10px 0 18px; }}
  .btns {{ margin: 8px 0 20px; }}
  .btn {{ display: inline-block; background: var(--navy); color: #fff !important;
         padding: 9px 16px; border-radius: 8px; font-size: 14px;
         margin-right: 10px; }}
  .btn:hover {{ opacity: .92; text-decoration: none; }}
  .meta {{ font-size: 13px; color: #555; margin-bottom: 10px; }}
  .tablewrap {{ overflow-x: auto; border: 1px solid var(--line);
               border-radius: 10px; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13.5px; }}
  thead th {{ position: sticky; top: 0; background: var(--band);
             color: var(--navy); text-align: left; padding: 10px 12px;
             white-space: nowrap; border-bottom: 1px solid var(--line); }}
  tbody td {{ padding: 9px 12px; border-bottom: 1px solid var(--line);
             vertical-align: middle; }}
  tbody tr:nth-child(even) {{ background: var(--alt); }}
  td.num {{ color: #888; width: 44px; }}
  td.brand {{ font-weight: 600; white-space: nowrap; }}
  td.c {{ white-space: nowrap; }}
  #q {{ padding: 9px 12px; border: 1px solid var(--line); border-radius: 8px;
       width: 280px; max-width: 100%; font-size: 14px; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="back"><a href="/">← All tools</a></div>
  <div class="kicker">Beauty</div>
  <h1>Sephora Brands</h1>
  <p class="desc">Every brand in Sephora's A–Z brand directory
    (<a href="{DATA_SOURCE_URL}" target="_blank" rel="noopener">sephora.com/brands-list</a>),
    each linking to its official Sephora brand page with auto-generated social
    handles. Download as CSV or Excel.</p>

  <div class="source">{DATA_SOURCE_LABEL} · {total} brands ({snapshot_note};
    {verified} verified, rest auto-generated)
    <a href="{DATA_SOURCE_URL}" target="_blank" rel="noopener">Source ↗</a></div>

  <p class="note"><strong>Note:</strong> Brand pages link to Sephora.
    Facebook / Instagram / X / YouTube / TikTok handles are auto-generated from
    the brand name and are <strong>unverified</strong> except for the
    {verified} brands flagged as verified. Refresh the full list any time by
    running <code>fetch_sephora_brands.py</code> locally.</p>

  <div class="btns">
    <a class="btn" href="/sephora-brands/export?fmt=csv">Download CSV</a>
    <a class="btn" href="/sephora-brands/export?fmt=xlsx">Download Excel</a>
  </div>

  <div class="meta">
    <input id="q" type="text" placeholder="Filter brands… (type to search)"
           oninput="filterRows(this.value)">
    <span id="count"></span>
  </div>

  <div class="tablewrap">
    <table id="tbl">
      <thead>
        <tr>
          <th>#</th><th>Brand</th><th>Sephora Page</th><th>Facebook</th>
          <th>Instagram</th><th>X / Twitter</th><th>YouTube</th>
          <th>TikTok</th><th>Wikipedia</th>
        </tr>
      </thead>
      <tbody>
{rows_html}
      </tbody>
    </table>
  </div>
</div>
<script>
  function filterRows(term) {{
    term = (term || "").trim().toLowerCase();
    var rows = document.querySelectorAll("#tbl tbody tr");
    var shown = 0;
    rows.forEach(function(tr) {{
      var name = tr.querySelector("td.brand").textContent.toLowerCase();
      var hit = !term || name.indexOf(term) !== -1;
      tr.style.display = hit ? "" : "none";
      if (hit) shown++;
    }});
    document.getElementById("count").textContent =
      term ? "  " + shown + " match" + (shown === 1 ? "" : "es") : "";
  }}
</script>
</body>
</html>"""


# --------------------------------------------------------------------------- #
# Routes
# --------------------------------------------------------------------------- #
@router.get("/sephora-brands", response_class=HTMLResponse)
def sephora_brands_page():
    return render_page()


@router.get("/sephora-brands/export")
def sephora_brands_export(fmt: str = Query("csv", pattern="^(csv|xlsx)$")):
    if fmt == "xlsx":
        data = export_xlsx()
        return StreamingResponse(
            io.BytesIO(data),
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": 'attachment; filename="sephora_brands.xlsx"'
            },
        )
    data = export_csv()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="sephora_brands.csv"'},
    )
