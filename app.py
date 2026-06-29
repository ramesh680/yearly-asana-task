"""
Yearly Asana task - a shared hub of yearly data tools.

The landing page lists tools; each tool has its own route. Add more tools by
appending to TOOLS and registering a new route + template.
"""
import datetime as _dt
import io

from flask import Flask, render_template, request, send_file, abort

from tools import best_hospitals, best_colleges, premier_league, saudi_pro_league, twitch_streamers, wnba_teams, motorsports, beauty_brands, nfl_teams, racquet_sports, golf_tours, nba_teams, nhl_teams, mls_teams, nwsl_teams, mlb_teams, milb_teams, brasileirao, bundesliga, laliga, serie_a, sp500, combat_sports, sporting_events, streaming_services, ligue1, vg_franchises, vg_platforms, vg_publishers, cpg_brands, leagues_revenue

app = Flask(__name__)

APP_NAME = "Yearly Asana task"

TOOLS = [
    {
        "key": "best-hospitals",
        "category": "Healthcare",
        "title": "Best Hospitals (US)",
        "description": (
            "Pulls the latest U.S. best-hospital lists from U.S. News & World "
            "Report (Honor Roll) and Newsweek/Statista (World's Best Hospitals), "
            "with each hospital's official website and social handles. View and "
            "download as CSV or Excel."
        ),
        "endpoint": "best_hospitals_view",
        "available": True,
    },
    {
        "key": "best-colleges",
        "category": "Education",
        "title": "Best Colleges (US)",
        "description": (
            "U.S. News Best National Universities (Top 100, 2026), with each "
            "school's official website and social handles (Facebook, Instagram, "
            "X, YouTube, LinkedIn). View and download as CSV or Excel."
        ),
        "endpoint": "best_colleges_view",
        "available": True,
    },
    {
        "key": "russell-3000",
        "category": "Finance",
        "title": "Russell 3000",
        "description": (
            "All Russell 3000 constituents (via iShares IWV holdings), ranked "
            "by index weight, with each company's official website, Wikipedia "
            "page, and social handles (Facebook, Instagram, X, YouTube, "
            "LinkedIn). View and download as CSV or Excel."
        ),
        "endpoint": "russell_3000_view",
        "available": True,
    },
    {
        "key": "premier-league",
        "category": "Sports",
        "title": "English Premier League",
        "description": (
            "The 20 English Premier League clubs (2025/26 season, via the "
            "official Premier League site), ordered by final league position, "
            "with each club's home city, stadium, points, official website and "
            "social handles (X, Instagram, Facebook, YouTube) plus Wikipedia. "
            "View and download as CSV or Excel."
        ),
        "endpoint": "premier_league_view",
        "available": True,
    },
    {
        "key": "saudi-pro-league",
        "category": "Sports",
        "title": "Saudi Pro League",
        "description": (
            "The 18 Saudi Pro League (Roshn Saudi League) clubs (2025/26 season, "
            "via the official Saudi Pro League site), ordered by final league "
            "position, with each club's home city, stadium, points, official "
            "website and social handles (X, Instagram, Facebook, YouTube) plus "
            "Wikipedia. View and download as CSV or Excel."
        ),
        "endpoint": "saudi_pro_league_view",
        "available": True,
    },
    {
        "key": "twitch-streamers",
        "category": "Streaming",
        "title": "Top Twitch Streamers",
        "description": (
            "The top channels from TwitchTracker's overall ranking (a 30-day "
            "blend of average viewers, followers, views and stream time), with "
            "each channel's average viewers, all-time peak, hours watched, "
            "Twitch link and social handles (X, YouTube) plus Wikipedia. View "
            "and download as CSV or Excel."
        ),
        "endpoint": "twitch_streamers_view",
        "available": True,
    },
    {
        "key": "wnba-teams",
        "category": "Basketball",
        "title": "WNBA Teams",
        "description": (
            "The 13 Women's National Basketball Association teams (2025 season, "
            "via wnba.com), ordered by final regular-season standing, with each "
            "team's home city, arena, win-loss record, official website and "
            "social handles (X, Instagram, Facebook) plus Wikipedia. View and "
            "download as CSV or Excel."
        ),
        "endpoint": "wnba_teams_view",
        "available": True,
    },
    {
        "key": "motorsports",
        "category": "Motorsport",
        "title": "Top Motorsports",
        "description": (
            "The 10 most popular motorsport series and events (Formula 1, "
            "NASCAR, MotoGP, IndyCar, Le Mans, WRC, Dakar, Formula E, MXGP, "
            "Isle of Man TT), with each one's discipline, official website and "
            "social handles (X, Instagram, YouTube) plus Wikipedia. Ranked by "
            "popularity. View and download as CSV or Excel."
        ),
        "endpoint": "motorsports_view",
        "available": True,
    },
    {
        "key": "beauty-brands",
        "category": "Beauty",
        "title": "Top Beauty Brands",
        "description": (
            "The top 500 beauty brands on Ulta.com. The top 10 are ranked by "
            "2025 online sales share (WWD/Navigo Marketing) with category, share, "
            "socials and Wikipedia; ranks 11-500 are Ulta's alphabetical brand "
            "directory, each linking to its official Ulta page. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "beauty_brands_view",
        "available": True,
    },
    {
        "key": "nfl-teams",
        "category": "American Football",
        "title": "National Football League",
        "description": (
            "All 32 NFL teams grouped by conference (AFC/NFC) and division, with "
            "each team's home city, stadium, official website and social handles "
            "(X, Instagram) plus Wikipedia. View and download as CSV or Excel."
        ),
        "endpoint": "nfl_teams_view",
        "available": True,
    },
    {
        "key": "racquet-sports",
        "category": "Racquet Sports",
        "title": "Racquet Sports",
        "description": (
            "Racket & racquet sports with their international governing body, "
            "official website and social handles (X, Instagram) plus Wikipedia. "
            "Major sports first, then others from Wikipedia's list. View and "
            "download as CSV or Excel."
        ),
        "endpoint": "racquet_sports_view",
        "available": True,
    },

    {
        "key": "golf-tours",
        "category": "Golf",
        "title": "Golf Tours",
        "description": (
            "Major professional golf tours and circuits. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "golf_tours_view",
        "available": True,
    },
    {
        "key": "nba-teams",
        "category": "Basketball",
        "title": "National Basketball Association",
        "description": (
            "All 30 NBA teams by conference and division. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "nba_teams_view",
        "available": True,
    },
    {
        "key": "nhl-teams",
        "category": "Ice Hockey",
        "title": "National Hockey League",
        "description": (
            "All 32 NHL teams by conference and division. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "nhl_teams_view",
        "available": True,
    },
    {
        "key": "mls-teams",
        "category": "Soccer",
        "title": "Major League Soccer",
        "description": (
            "All 30 MLS clubs by conference. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "mls_teams_view",
        "available": True,
    },
    {
        "key": "nwsl-teams",
        "category": "Soccer",
        "title": "National Women's Soccer League",
        "description": (
            "NWSL clubs with home city and stadium. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "nwsl_teams_view",
        "available": True,
    },

    {
        "key": "mlb-teams",
        "category": "Baseball",
        "title": "Major League Baseball",
        "description": (
            "All 30 MLB teams by league and division. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "mlb_teams_view",
        "available": True,
    },
    {
        "key": "milb-teams",
        "category": "Baseball",
        "title": "Minor League Baseball",
        "description": (
            "Triple-A clubs (top minor-league tier) with their MLB parent club. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "milb_teams_view",
        "available": True,
    },
    {
        "key": "brasileirao",
        "category": "Soccer",
        "title": "Brasileirão Série A",
        "description": (
            "Campeonato Brasileiro Série A clubs (2025 season). With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "brasileirao_view",
        "available": True,
    },
    {
        "key": "bundesliga",
        "category": "Soccer",
        "title": "Bundesliga",
        "description": (
            "Bundesliga clubs (2025-26 season). With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "bundesliga_view",
        "available": True,
    },
    {
        "key": "laliga",
        "category": "Soccer",
        "title": "LaLiga",
        "description": (
            "LaLiga clubs (2025-26 season). With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "laliga_view",
        "available": True,
    },

    {
        "key": "serie-a",
        "category": "Soccer",
        "title": "Serie A",
        "description": (
            "Serie A clubs (2025-26 season). With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "serie_a_view",
        "available": True,
    },
    {
        "key": "sp500",
        "category": "Finance",
        "title": "S&P 500",
        "description": (
            "S&P 500 index constituents (503 securities). With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "sp500_view",
        "available": True,
    },
    {
        "key": "combat-sports",
        "category": "Combat Sports",
        "title": "Combat Sports",
        "description": (
            "Major combat sports and disciplines with their governing body. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "combat_sports_view",
        "available": True,
    },
    {
        "key": "sporting-events",
        "category": "Sports Events",
        "title": "Sporting Events",
        "description": (
            "The biggest global sporting events, ranked by popularity. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "sporting_events_view",
        "available": True,
    },
    {
        "key": "streaming-services",
        "category": "Streaming",
        "title": "Streaming Services",
        "description": (
            "Notable streaming media services (video, music, live, audio). With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "streaming_services_view",
        "available": True,
    },

    {
        "key": "ligue1",
        "category": "Soccer",
        "title": "Ligue 1",
        "description": (
            "Ligue 1 clubs (2025-26 season). With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "ligue1_view",
        "available": True,
    },
    {
        "key": "vg-franchises",
        "category": "Video Games",
        "title": "Video Game Franchises",
        "description": (
            "Major video game franchises with their publisher. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "vg_franchises_view",
        "available": True,
    },
    {
        "key": "vg-platforms",
        "category": "Video Games",
        "title": "Video Game Platforms",
        "description": (
            "Major video game consoles and platforms with their manufacturer. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "vg_platforms_view",
        "available": True,
    },
    {
        "key": "vg-publishers",
        "category": "Video Games",
        "title": "Video Game Publishers",
        "description": (
            "Major video game publishers. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "vg_publishers_view",
        "available": True,
    },
    {
        "key": "cpg-brands",
        "category": "Consumer Goods",
        "title": "CPG Brands",
        "description": (
            "Major consumer packaged goods (CPG) brands with their parent company. With official website and social handles (Facebook, "
            "Instagram, X, YouTube, TikTok) plus Wikipedia. View and download "
            "as CSV or Excel."
        ),
        "endpoint": "cpg_brands_view",
        "available": True,
    },

    {
        "key": "leagues-revenue",
        "category": "Sports Business",
        "title": "Sports Leagues by Revenue",
        "description": (
            "Professional sports leagues ranked by annual revenue (Wikipedia), "
            "with official website and social handles (Facebook, Instagram, X, "
            "YouTube, TikTok) plus Wikipedia. View and download as CSV or Excel."
        ),
        "endpoint": "leagues_revenue_view",
        "available": True,
    },
]


@app.context_processor
def inject_globals():
    return {"app_name": APP_NAME, "today": _dt.date.today().isoformat()}


@app.route("/")
def index():
    return render_template("index.html", tools=TOOLS)


# ---------------------------------------------------------------- Hospitals
@app.route("/best-hospitals")
def best_hospitals_view():
    source = request.args.get("source", "newsweek")
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = best_hospitals.get_hospitals(source, live=live)
    return render_template(
        "best_hospitals.html", rows=rows, meta=meta, sources=best_hospitals.SOURCES,
    )


@app.route("/best-hospitals/export")
def best_hospitals_export():
    source = request.args.get("source", "newsweek")
    fmt = request.args.get("fmt", "csv")
    rows, meta = best_hospitals.get_hospitals(source)
    stamp = _dt.date.today().isoformat()
    base = "best_hospitals_{}_{}_{}".format(meta["source"], meta["edition"], stamp)
    if fmt == "xlsx":
        return _send_xlsx(best_hospitals.columns(meta), rows, "Best Hospitals", base + ".xlsx")
    return _send_csv(best_hospitals.to_csv(rows, meta), base + ".csv")


# ---------------------------------------------------------------- Colleges
@app.route("/best-colleges")
def best_colleges_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = best_colleges.get_colleges(live=live)
    return render_template("best_colleges.html", rows=rows, meta=meta)


@app.route("/best-colleges/export")
def best_colleges_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = best_colleges.get_colleges()
    stamp = _dt.date.today().isoformat()
    base = "best_colleges_national_universities_{}_{}".format(meta["edition"], stamp)
    if fmt == "xlsx":
        return _send_xlsx(best_colleges.columns(), rows, "Best Colleges", base + ".xlsx")
    return _send_csv(best_colleges.to_csv(rows), base + ".csv")

# ---------------------------------------------------------------- Russell 3000
@app.route("/russell-3000")
def russell_3000_view():
    import os
    path = os.path.join(app.root_path, "templates", "russell-3000.html")
    with open(path, encoding="utf-8") as f:
        return f.read()

# ------------------------------------------------------------ Premier League
@app.route("/premier-league")
def premier_league_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = premier_league.get_clubs(live=live)
    return render_template("premier-league.html", rows=rows, meta=meta)


@app.route("/premier-league/export")
def premier_league_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = premier_league.get_clubs()
    stamp = _dt.date.today().isoformat()
    base = "premier_league_clubs_{}_{}".format(meta["edition"].replace("/", "-"), stamp)
    if fmt == "xlsx":
        return _send_xlsx(premier_league.columns(), rows, "Premier League", base + ".xlsx")
    return _send_csv(premier_league.to_csv(rows), base + ".csv")


# ---------------------------------------------------------- - Saudi Pro League
@app.route("/saudi-pro-league")
def saudi_pro_league_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = saudi_pro_league.get_clubs(live=live)
    return render_template("saudi-pro-league.html", rows=rows, meta=meta)


@app.route("/saudi-pro-league/export")
def saudi_pro_league_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = saudi_pro_league.get_clubs()
    stamp = _dt.date.today().isoformat()
    base = "saudi_pro_league_clubs_{}_{}".format(meta["edition"].replace("/", "-"), stamp)
    if fmt == "xlsx":
        return _send_xlsx(saudi_pro_league.columns(), rows, "Saudi Pro League", base + ".xlsx")
    return _send_csv(saudi_pro_league.to_csv(rows), base + ".csv")


# ----------------------------------------------------------- Twitch Streamers
@app.route("/twitch-streamers")
def twitch_streamers_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = twitch_streamers.get_streamers(live=live)
    return render_template("twitch-streamers.html", rows=rows, meta=meta)


@app.route("/twitch-streamers/export")
def twitch_streamers_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = twitch_streamers.get_streamers()
    stamp = _dt.date.today().isoformat()
    base = "top_twitch_streamers_{}".format(stamp)
    if fmt == "xlsx":
        return _send_xlsx(twitch_streamers.columns(), rows, "Top Twitch Streamers", base + ".xlsx")
    return _send_csv(twitch_streamers.to_csv(rows), base + ".csv")


# ---------------------------------------------------------------- WNBA Teams
@app.route("/wnba-teams")
def wnba_teams_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = wnba_teams.get_teams(live=live)
    return render_template("wnba-teams.html", rows=rows, meta=meta)


@app.route("/wnba-teams/export")
def wnba_teams_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = wnba_teams.get_teams()
    stamp = _dt.date.today().isoformat()
    base = "wnba_teams_{}_{}".format(meta["edition"].replace(" ", "_"), stamp)
    if fmt == "xlsx":
        return _send_xlsx(wnba_teams.columns(), rows, "WNBA Teams", base + ".xlsx")
    return _send_csv(wnba_teams.to_csv(rows), base + ".csv")


# ---------------------------------------------------------------- Motorsports
@app.route("/motorsports")
def motorsports_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = motorsports.get_motorsports(live=live)
    return render_template("motorsports.html", rows=rows, meta=meta)


@app.route("/motorsports/export")
def motorsports_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = motorsports.get_motorsports()
    stamp = _dt.date.today().isoformat()
    base = "top_motorsports_{}".format(stamp)
    if fmt == "xlsx":
        return _send_xlsx(motorsports.columns(), rows, "Top Motorsports", base + ".xlsx")
    return _send_csv(motorsports.to_csv(rows), base + ".csv")


# ------------------------------------------------------------- Beauty Brands
@app.route("/beauty-brands")
def beauty_brands_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = beauty_brands.get_brands(live=live)
    return render_template("beauty-brands.html", rows=rows, meta=meta)


@app.route("/beauty-brands/export")
def beauty_brands_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = beauty_brands.get_brands()
    stamp = _dt.date.today().isoformat()
    base = "top_beauty_brands_ulta_{}".format(stamp)
    if fmt == "xlsx":
        return _send_xlsx(beauty_brands.columns(), rows, "Top Beauty Brands", base + ".xlsx")
    return _send_csv(beauty_brands.to_csv(rows), base + ".csv")


# ----------------------------------------------------------- National Football League
@app.route("/nfl-teams")
def nfl_teams_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = nfl_teams.get_teams(live=live)
    return render_template("nfl-teams.html", rows=rows, meta=meta)


@app.route("/nfl-teams/export")
def nfl_teams_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = nfl_teams.get_teams()
    stamp = _dt.date.today().isoformat()
    base = "nfl_teams_{}".format(stamp)
    if fmt == "xlsx":
        return _send_xlsx(nfl_teams.columns(), rows, "NFL Teams", base + ".xlsx")
    return _send_csv(nfl_teams.to_csv(rows), base + ".csv")


# ----------------------------------------------------------- Racquet Sports
@app.route("/racquet-sports")
def racquet_sports_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = racquet_sports.get_sports(live=live)
    return render_template("racquet-sports.html", rows=rows, meta=meta)


@app.route("/racquet-sports/export")
def racquet_sports_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = racquet_sports.get_sports()
    stamp = _dt.date.today().isoformat()
    base = "racquet_sports_{}".format(stamp)
    if fmt == "xlsx":
        return _send_xlsx(racquet_sports.columns(), rows, "Racquet Sports", base + ".xlsx")
    return _send_csv(racquet_sports.to_csv(rows), base + ".csv")


@app.route("/golf-tours")
def golf_tours_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = golf_tours.get_rows(live=live)
    return render_template("golf-tours.html", rows=rows, meta=meta)


@app.route("/golf-tours/export")
def golf_tours_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = golf_tours.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "golf_tours_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(golf_tours.columns(), rows, "Golf Tours", base + ".xlsx")
    return _send_csv(golf_tours.to_csv(rows), base + ".csv")


@app.route("/nba-teams")
def nba_teams_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = nba_teams.get_rows(live=live)
    return render_template("nba-teams.html", rows=rows, meta=meta)


@app.route("/nba-teams/export")
def nba_teams_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = nba_teams.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "nba_teams_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(nba_teams.columns(), rows, "National Basketball Association", base + ".xlsx")
    return _send_csv(nba_teams.to_csv(rows), base + ".csv")


@app.route("/nhl-teams")
def nhl_teams_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = nhl_teams.get_rows(live=live)
    return render_template("nhl-teams.html", rows=rows, meta=meta)


@app.route("/nhl-teams/export")
def nhl_teams_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = nhl_teams.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "nhl_teams_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(nhl_teams.columns(), rows, "National Hockey League", base + ".xlsx")
    return _send_csv(nhl_teams.to_csv(rows), base + ".csv")


@app.route("/mls-teams")
def mls_teams_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = mls_teams.get_rows(live=live)
    return render_template("mls-teams.html", rows=rows, meta=meta)


@app.route("/mls-teams/export")
def mls_teams_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = mls_teams.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "mls_teams_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(mls_teams.columns(), rows, "Major League Soccer", base + ".xlsx")
    return _send_csv(mls_teams.to_csv(rows), base + ".csv")


@app.route("/nwsl-teams")
def nwsl_teams_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = nwsl_teams.get_rows(live=live)
    return render_template("nwsl-teams.html", rows=rows, meta=meta)


@app.route("/nwsl-teams/export")
def nwsl_teams_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = nwsl_teams.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "nwsl_teams_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(nwsl_teams.columns(), rows, "National Women's Soccer League", base + ".xlsx")
    return _send_csv(nwsl_teams.to_csv(rows), base + ".csv")


@app.route("/mlb-teams")
def mlb_teams_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = mlb_teams.get_rows(live=live)
    return render_template("mlb-teams.html", rows=rows, meta=meta)


@app.route("/mlb-teams/export")
def mlb_teams_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = mlb_teams.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "mlb_teams_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(mlb_teams.columns(), rows, "Major League Baseball", base + ".xlsx")
    return _send_csv(mlb_teams.to_csv(rows), base + ".csv")


@app.route("/milb-teams")
def milb_teams_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = milb_teams.get_rows(live=live)
    return render_template("milb-teams.html", rows=rows, meta=meta)


@app.route("/milb-teams/export")
def milb_teams_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = milb_teams.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "milb_teams_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(milb_teams.columns(), rows, "Minor League Baseball", base + ".xlsx")
    return _send_csv(milb_teams.to_csv(rows), base + ".csv")


@app.route("/brasileirao")
def brasileirao_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = brasileirao.get_rows(live=live)
    return render_template("brasileirao.html", rows=rows, meta=meta)


@app.route("/brasileirao/export")
def brasileirao_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = brasileirao.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "brasileirao_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(brasileirao.columns(), rows, "Brasileirão Série A", base + ".xlsx")
    return _send_csv(brasileirao.to_csv(rows), base + ".csv")


@app.route("/bundesliga")
def bundesliga_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = bundesliga.get_rows(live=live)
    return render_template("bundesliga.html", rows=rows, meta=meta)


@app.route("/bundesliga/export")
def bundesliga_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = bundesliga.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "bundesliga_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(bundesliga.columns(), rows, "Bundesliga", base + ".xlsx")
    return _send_csv(bundesliga.to_csv(rows), base + ".csv")


@app.route("/laliga")
def laliga_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = laliga.get_rows(live=live)
    return render_template("laliga.html", rows=rows, meta=meta)


@app.route("/laliga/export")
def laliga_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = laliga.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "laliga_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(laliga.columns(), rows, "LaLiga", base + ".xlsx")
    return _send_csv(laliga.to_csv(rows), base + ".csv")


@app.route("/serie-a")
def serie_a_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = serie_a.get_rows(live=live)
    return render_template("serie-a.html", rows=rows, meta=meta)


@app.route("/serie-a/export")
def serie_a_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = serie_a.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "serie_a_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(serie_a.columns(), rows, "Serie A", base + ".xlsx")
    return _send_csv(serie_a.to_csv(rows), base + ".csv")


@app.route("/sp500")
def sp500_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = sp500.get_rows(live=live)
    return render_template("sp500.html", rows=rows, meta=meta)


@app.route("/sp500/export")
def sp500_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = sp500.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "sp500_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(sp500.columns(), rows, "S&P 500", base + ".xlsx")
    return _send_csv(sp500.to_csv(rows), base + ".csv")


@app.route("/combat-sports")
def combat_sports_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = combat_sports.get_rows(live=live)
    return render_template("combat-sports.html", rows=rows, meta=meta)


@app.route("/combat-sports/export")
def combat_sports_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = combat_sports.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "combat_sports_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(combat_sports.columns(), rows, "Combat Sports", base + ".xlsx")
    return _send_csv(combat_sports.to_csv(rows), base + ".csv")


@app.route("/sporting-events")
def sporting_events_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = sporting_events.get_rows(live=live)
    return render_template("sporting-events.html", rows=rows, meta=meta)


@app.route("/sporting-events/export")
def sporting_events_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = sporting_events.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "sporting_events_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(sporting_events.columns(), rows, "Sporting Events", base + ".xlsx")
    return _send_csv(sporting_events.to_csv(rows), base + ".csv")


@app.route("/streaming-services")
def streaming_services_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = streaming_services.get_rows(live=live)
    return render_template("streaming-services.html", rows=rows, meta=meta)


@app.route("/streaming-services/export")
def streaming_services_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = streaming_services.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "streaming_services_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(streaming_services.columns(), rows, "Streaming Services", base + ".xlsx")
    return _send_csv(streaming_services.to_csv(rows), base + ".csv")


@app.route("/ligue1")
def ligue1_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = ligue1.get_rows(live=live)
    return render_template("ligue1.html", rows=rows, meta=meta)


@app.route("/ligue1/export")
def ligue1_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = ligue1.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "ligue1_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(ligue1.columns(), rows, "Ligue 1", base + ".xlsx")
    return _send_csv(ligue1.to_csv(rows), base + ".csv")


@app.route("/vg-franchises")
def vg_franchises_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = vg_franchises.get_rows(live=live)
    return render_template("vg-franchises.html", rows=rows, meta=meta)


@app.route("/vg-franchises/export")
def vg_franchises_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = vg_franchises.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "vg_franchises_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(vg_franchises.columns(), rows, "Video Game Franchises", base + ".xlsx")
    return _send_csv(vg_franchises.to_csv(rows), base + ".csv")


@app.route("/vg-platforms")
def vg_platforms_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = vg_platforms.get_rows(live=live)
    return render_template("vg-platforms.html", rows=rows, meta=meta)


@app.route("/vg-platforms/export")
def vg_platforms_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = vg_platforms.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "vg_platforms_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(vg_platforms.columns(), rows, "Video Game Platforms", base + ".xlsx")
    return _send_csv(vg_platforms.to_csv(rows), base + ".csv")


@app.route("/vg-publishers")
def vg_publishers_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = vg_publishers.get_rows(live=live)
    return render_template("vg-publishers.html", rows=rows, meta=meta)


@app.route("/vg-publishers/export")
def vg_publishers_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = vg_publishers.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "vg_publishers_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(vg_publishers.columns(), rows, "Video Game Publishers", base + ".xlsx")
    return _send_csv(vg_publishers.to_csv(rows), base + ".csv")


@app.route("/cpg-brands")
def cpg_brands_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = cpg_brands.get_rows(live=live)
    return render_template("cpg-brands.html", rows=rows, meta=meta)


@app.route("/cpg-brands/export")
def cpg_brands_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = cpg_brands.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "cpg_brands_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(cpg_brands.columns(), rows, "CPG Brands", base + ".xlsx")
    return _send_csv(cpg_brands.to_csv(rows), base + ".csv")


@app.route("/leagues-revenue")
def leagues_revenue_view():
    live = request.args.get("live") in ("1", "true", "yes")
    rows, meta = leagues_revenue.get_rows(live=live)
    return render_template("leagues-revenue.html", rows=rows, meta=meta)


@app.route("/leagues-revenue/export")
def leagues_revenue_export():
    fmt = request.args.get("fmt", "csv")
    rows, meta = leagues_revenue.get_rows()
    stamp = _dt.date.today().isoformat()
    base = "leagues_by_revenue_" + stamp
    if fmt == "xlsx":
        return _send_xlsx(leagues_revenue.columns(), rows, "Sports Leagues by Revenue", base + ".xlsx")
    return _send_csv(leagues_revenue.to_csv(rows), base + ".csv")


# ---------------------------------------------------------------- Helpers
def _send_csv(csv_text, filename):
    return send_file(
        io.BytesIO(csv_text.encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


def _send_xlsx(cols, rows, sheet_title, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except Exception:
        abort(500, "openpyxl is not installed; CSV export is still available.")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append([label for label, _key in cols])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(key, "") for _label, key in cols])

    width_by_label = {
        "Team": 24, "Conference": 12, "Division": 10, "Stadium": 30, "Sport": 22, "Governing Body": 40,
        "Rank": 8, "Hospital": 52, "University": 44, "City": 16, "State": 18,
        "Score": 9, "Website": 30, "Facebook": 42, "Instagram": 38,
        "X / Twitter": 30, "YouTube": 44, "LinkedIn": 52, "Wikipedia": 30,
        "Position": 9, "Club": 30, "Stadium": 28, "Points": 9,
        "Channel": 26, "Avg Viewers": 14, "Peak Viewers": 14,
        "Hours Watched": 16, "Twitch": 28,
        "Seed": 7, "Team": 26, "Arena": 26, "W": 6, "L": 6,
        "Series / Event": 26, "Category": 26,
        "Brand": 24, "Ulta.com Share": 14,
    }
    for i, (label, _key) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width_by_label.get(label, 24)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    import os

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
